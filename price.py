import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, numbers
import logging
from datetime import datetime
import shutil

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

BIBLE_SPREADSHEET_ID = "1QB3Jv7cL5hNwDKx9rQF6FCrKHW7IHPAqrUg7FIvY7Dk"

def get_sheets_service():
    try:
        from google.oauth2.service_account import Credentials
        credentials = Credentials.from_service_account_file(
            os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
        )
        service = build('sheets', 'v4', credentials=credentials)
        return service
    except Exception as e:
        logger.error(f"Ошибка инициализации Google Sheets API: {e}")
        raise

def load_bible_data():
    """
    Загружает данные из Google Sheets таблицы Bible.xlsx и возвращает их в виде DataFrame.
    Ожидается, что данные находятся на листе с именем "Bible" и в диапазоне A2:D,
    где строка 1 содержит заголовки: FAQ, Answers, Verification, rule.
    """
    try:
        service = get_sheets_service()
        range_name = "Bible!A2:D"
        result = service.spreadsheets().values().get(
            spreadsheetId=BIBLE_SPREADSHEET_ID,
            range=range_name
        ).execute()
        values = result.get("values", [])
        if values:
            df = pd.DataFrame(values, columns=["FAQ", "Answers", "Verification", "rule"])
        else:
            df = pd.DataFrame(columns=["FAQ", "Answers", "Verification", "rule"])
        logger.info(f"Bible data loaded. Количество записей: {len(df)}")
        return df
    except Exception as e:
        logger.error(f"Ошибка при загрузке данных из Bible.xlsx: {e}")
        return None

def upload_or_update_file(file_name, file_stream):
    pass

def ensure_local_bible_file(local_path):
    if not os.path.exists(local_path):
        try:
            directory = os.path.dirname(local_path)
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.append(["FAQ", "Answers", "Verification", "rule"])
            wb.save(local_path)
            logger.info(f"Локальный файл {local_path} создан с заголовками.")
        except Exception as e:
            logger.error(f"Ошибка при создании локального файла {local_path}: {e}")
            raise

def save_bible_pair(question, answer):
    try:
        service = get_sheets_service()
        new_row = [[question, answer, "Check", ""]]
        body = {"values": new_row}
        result = service.spreadsheets().values().append(
            spreadsheetId=BIBLE_SPREADSHEET_ID,
            range="Bible!A:D",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body=body
        ).execute()
        logger.info(f"Новая пара добавлена в Google Sheets: FAQ='{question}', Answers='{answer}', Verification='Check'. Ответ API: {result}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении пары в Google Sheets: {e}")
        try:
            temp_file = os.path.join(os.getcwd(), "Temp_Bible.xlsx")
            ensure_local_bible_file(temp_file)
            wb = load_workbook(temp_file)
            ws = wb.active
            ws.append([question, answer, "Check", ""])
            wb.save(temp_file)
            logger.error(f"Временный файл {temp_file} создан из-за ошибки записи в оригинальный файл.")
        except Exception as e2:
            logger.error(f"Ошибка при создании временного файла Temp_Bible.xlsx: {e2}")
        raise

    try:
        today_str = datetime.now().strftime("%Y%m%d")
        backup_file = os.path.join(os.getcwd(), "CAEC_API_Data", "BIG_DATA", f"Reserv_Bible_{today_str}.xlsx")
        if not os.path.exists(backup_file):
            df = load_bible_data()
            if df is not None:
                df.to_excel(backup_file, index=False)
                logger.info(f"Резервная копия создана: {backup_file}")
    except Exception as e:
        logger.error(f"Ошибка при создании резервной копии Reserv_Bible: {e}")
