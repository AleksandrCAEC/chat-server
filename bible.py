import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from datetime import datetime
import shutil
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Укажите идентификатор Google Sheets таблицы Bible.xlsx
BIBLE_SPREADSHEET_ID = "YOUR_BIBLE_SPREADSHEET_ID"

def get_sheets_service():
    try:
        credentials = Credentials.from_service_account_file(
            os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
        )
        service = build('sheets', 'v4', credentials=credentials)
        return service
    except Exception as e:
        logger.error(f"Error initializing Sheets API: {e}")
        raise

def load_bible_data():
    """
    Загружает данные из Google Sheets таблицы Bible.xlsx.
    Ожидается, что данные находятся на листе "Bible" в диапазоне A2:D,
    где строка 1 содержит заголовки: FAQ, Answers, Verification, rule.
    """
    try:
        service = get_sheets_service()
        range_name = "Bible!A2:D"
        result = service.spreadsheets().values().get(
            spreadsheetId=BIBLE_SPREADSHEET_ID,
            range=range_name
        ).execute()
        values = result.get("values", [])
        if values:
            df = pd.DataFrame(values, columns=["FAQ", "Answers", "Verification", "rule"])
        else:
            df = pd.DataFrame(columns=["FAQ", "Answers", "Verification", "rule"])
        logger.info(f"Bible data loaded. Records: {len(df)}")
        return df
    except Exception as e:
        logger.error(f"Error loading Bible data: {e}")
        return None

def save_bible_pair(question, answer):
    """
    Добавляет новую строку в таблицу Bible.xlsx.
    """
    try:
        service = get_sheets_service()
        new_row = [[question, answer, "Check", ""]]
        body = {"values": new_row}
        result = service.spreadsheets().values().append(
            spreadsheetId=BIBLE_SPREADSHEET_ID,
            range="Bible!A:D",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body=body
        ).execute()
        logger.info(f"New pair added: FAQ='{question}', Answers='{answer}', Verification='Check'. API response: {result}")
    except Exception as e:
        logger.error(f"Error saving pair to Bible: {e}")
        # Попытка сохранить во временный файл
        try:
            temp_file = os.path.join(os.getcwd(), "Temp_Bible.xlsx")
            ensure_local_bible_file(temp_file)
            wb = load_workbook(temp_file)
            ws = wb.active
            ws.append([question, answer, "Check", ""])
            wb.save(temp_file)
            logger.error(f"Temporary Bible file created: {temp_file}")
        except Exception as e2:
            logger.error(f"Error creating temporary Bible file: {e2}")
        raise

def ensure_local_bible_file(local_path):
    """
    Создает локальный файл Bible.xlsx, если он не существует.
    """
    if not os.path.exists(local_path):
        try:
            directory = os.path.dirname(local_path)
            if not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.append(["FAQ", "Answers", "Verification", "rule"])
            wb.save(local_path)
            logger.info(f"Local Bible file created: {local_path}")
        except Exception as e:
            logger.error(f"Error creating local Bible file {local_path}: {e}")
            raise

def get_rule(rule_key):
    """
    Возвращает правило по заданному ключу.
    В данной реализации возвращается заглушка.
    В реальной системе можно реализовать поиск строки с FAQ равным rule_key.
    """
    return f"<{rule_key}>"

if __name__ == "__main__":
    df = load_bible_data()
    logger.info(df)
